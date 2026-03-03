# AI툴 (Gemini 기반 웹사이트 자동화 검색)
import google.generativeai as genai
import requests
from bs4 import BeautifulSoup
import pandas as pd
import json
import time # [추가] 대기 시간을 만들기 위해 필요합니다.
from datetime import datetime, timedelta
# [수정] 심화형 사이트 처리를 위한 라이브러리 추가
from playwright.sync_api import sync_playwright
# [추가] 엑셀에 색상을 입히기 위한 라이브러리
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# [설정] 발급받으신 Gemini API 키를 입력하세요
genai.configure(api_key="AIzaSyDmiPZdmd37F7Sy_hcuRAmgoiwCXetfr2I")
model = genai.GenerativeModel('gemini-2.5-flash')

# --------------------------------------일반형 추출 로직-------------------------------------
def get_web_text_normal(url): #일반형 추출
    """일반형 사이트 전용 추출 함수 (기존 로직)"""
    try:
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
        print(f"📡 사이트 접속 중...") # [추가 팁]
        response = requests.get(url, headers=headers, timeout=20)
        response.encoding = 'utf-8'
        soup = BeautifulSoup(response.text, 'html.parser')
        # 불필요한 태그 제거
        for script in soup(["script", "style", "nav", "footer", "header", "aside"]):
            script.extract()
        # [수정] 데이터 양을 줄이기 위해 게시판 영역(table)만 먼저 찾습니다.
        board_content = soup.find('table')
        if board_content:
            return board_content.get_text()
        return soup.get_text()
    except Exception as e:
        return f"Error: {e}"
        
# --------------------------------------심화형 추출 로직-------------------------------------
def get_web_text_advanced(url):
    """[심화] 심화형 사이트 전용 추출 함수 (브라우저 실행 방식)"""
    try:
        with sync_playwright() as p:
            print(f"🖥️ 가상 브라우저 실행 중...")
            # 브라우저 열기 (headless=True는 창을 띄우지 않고 백그라운드에서 실행)
            browser = p.chromium.launch(headless=True)
            page = browser.new_page()
            
            # 페이지 접속 및 데이터 로딩 대기
            page.goto(url, wait_until="domcontentloaded", timeout=60000)
            print(f"⏳ 데이터 로딩 대기 중 (7초)...")
            time.sleep(7) # 나라장터 등 동적 데이터 로딩을 위한 추가 대기
            
            # 전체 페이지 소스 가져오기
            html_content = page.content()
            browser.close()
            
            # BeautifulSoup으로 텍스트만 정제
            soup = BeautifulSoup(html_content, 'html.parser')
            for extra in soup(["script", "style", "nav", "footer", "header", "aside"]):
                extra.extract()
            
            # 데이터가 있는 본문 영역만 좁히기 (나라장터는 보통 table 안에 있음)
            return soup.get_text(separator='\n', strip=True)
    except Exception as e:
        return f"Error: {e}"

def ask_gemini_smart_parse(text, org_name):
       
    # # 날짜기준 검색 로직
    today = datetime.now() #현재날짜 기준
    three_days_ago = today - timedelta(days=3) #날짜지정 현재날짜 기준 3일이전까지 검색
    
    # AI에게 줄 명령문(Prompt)
    prompt = f"""
    너는 관공서 정보 분석가야. 제시된 텍스트는 웹사이트에서 추출된 로우 데이터(Raw Data)야. 
    이 안에는 게시판 형태의 목록 데이터가 포함되어 있어.

   [지침]
    1. 기관명 지정: 결과의 '기관명' 필드에는 반드시 "{org_name}"라고 적어줘.
    2. 기간: 게시날짜가 {three_days_ago.strftime('%Y-%m-%d')} 이후인 모든 공고를 추출해 (키워드 상관없음).
    3. 점검판별: 제목에 "점검", "수행기관", "지정", "안전" 등의 단어가 포함되어 있으면 '확인여부' 필드에 "Y", 아니면 "N"이라고 적어줘.
    4. 형식: 아래 한글 필드명을 가진 JSON 리스트로만 응답해.

    [응답 항목]
    - 기관명 
    - 제목 (공고 제목)
    - 게시날짜 (YYYY-MM-DD)
    - 작성자 (부서 또는 성함)
    - 점검여부 (Yes or No)
    
    [분석할 내용]:
    {text[:12000]}  # 토큰 제한을 고려해 앞부분 1만자만 전송
    """
    
    response = model.generate_content(prompt)
    
    # 응답 정제
    try:
        print(f"🧠 AI 분석 중... (최근 3일 이내 모든 글 추출 중)")
        response = model.generate_content(prompt)
        # JSON 문자열만 추출하기 위한 처리 추가
        res_text = response.text.strip()
        # [보완] AI가 빈 응답을 하거나 이상한 말을 할 경우 예외 처리
        if not res_text or len(res_text) < 2:
            print(f"ℹ️ AI로부터 빈 응답을 받았습니다. (데이터 없음으로 간주)")
            return []
        if "```json" in res_text:
            res_text = res_text.split("```json")[1].split("```")[0].strip()
        elif "```" in res_text:
            res_text = res_text.split("```")[1].split("```")[0].strip()
        raw_res = res_text
        return json.loads(raw_res)
    except Exception as e:
        # [수정] 할당량 초과 에러 시 안내 문구 출력
        if "429" in str(e):
            print("⚠️ API 요청 제한에 도달했습니다.")
        else:
            print(f"⚠️ {org_name} 분석 중 데이터 해석 오류: {e}")
            # print(f"--- AI 응답 원본 ---\n{response.text}\n------------------") # 필요시 주석 해제하여 확인
        return []

# --- 실행부 ---
try:
    site_df = pd.read_excel("site_list.xlsx")
    target_sites = site_df.to_dict('records')
except Exception as e:
    print(f"❌ 파일을 읽을 수 없습니다: {e}")
    target_sites = []

all_collected_data = []

for site in target_sites:
    url = site['URL']
    org_name = site['기관명']
    print(f"\n🌐 분석 진행 중: {org_name}")
    
    # [수정] 타입에 따른 추출 함수 분기
    if site.get('타입') == '심화':
        web_text = get_web_text_advanced(url)
    else:
        web_text = get_web_text_normal(url)
    
    if "Error" not in web_text:
        results = ask_gemini_smart_parse(web_text, org_name)
        if results:
            all_collected_data.extend(results) # [수정] 결과 데이터 추가
            print(f"✅ {org_name}에서 {len(results)}건 발견")
        else:
            print(f"ℹ️ 최근 3일 이내 신규공고 없음.")

# [수정] API 할당량 관리를 위해 사이트당 5초간 휴식 (무료 티어 필수)
    # print(f"⏱️ 다음 사이트 분석을 위해 1초간 대기합니다...")
    time.sleep(1)

# 결과 출력 및 저장 (제시해주신 구조 유지)
if all_collected_data:
    df = pd.DataFrame(all_collected_data)
    # [수정] 중복 제거 기준을 '기관명'과 '제목' 모두 확인하도록 보완
    if '제목' in df.columns and '기관명' in df.columns:
        df = df.drop_duplicates(subset=['기관명', '제목'])
    # 엑셀 파일 생성
    file_name = f"점검항목_수집결과_{datetime.now().strftime('%Y%m%d')}.xlsx"
    df.to_excel(file_name, index=False)
    # [추가] 엑셀 강조 표시 로직
    try:
        wb = load_workbook(file_name)
        ws = wb.active
        # 노란색 채우기 설정
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # '점검여부' 열(보통 5번째 열)이 "Y"인 행에 색상 칠하기
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=5).value == "Y":
                for col in range(1, 6): # 1번부터 5번 열까지
                    ws.cell(row=row, column=col).fill = yellow_fill
        
        wb.save(file_name)
        print(f"\n✨ 모든 작업 완료! {file_name}을 확인하세요.")
        print(f"💡 노란색으로 강조된 행이 '점검' 검색 공고입니다.")
    except Exception as e:
        print(f"⚠️ 엑셀 강조 작업 중 오류: {e}")
    print(f"총 {len(df)}건의 신규 공고를 찾았습니다.")
    print(f"파일명: {file_name}")
else:
    print("\n❌ 모든사이트 확인결과 3일 이내의 공고가 없습니다.")